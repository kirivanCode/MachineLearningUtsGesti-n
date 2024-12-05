import random
import streamlit as st
import pandas as pd
import numpy as np
from sklearn.neighbors import KNeighborsClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.metrics import confusion_matrix, classification_report
import requests
from faker import Faker
import joblib
import os
from datetime import datetime, time, timedelta
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import openpyxl
from streamlit_option_menu import option_menu
import json


if not os.path.exists('models'):
    os.makedirs('models')

# Configuración inicial donde hago la conexion con la api zzzzz
fake = Faker()
BASE_URL = "http://localhost:8000/api"


# Configuración de la página
st.set_page_config(
    page_title="Gestión UTS",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Aplicar styles, es como el css
st.markdown("""
    <style>
    .stButton>button {
        background-color: #0066cc;
        color: white;
        border-radius: 5px;
        padding: 0.5rem 1rem;
        border: none;
    }
    .stButton>button:hover {
        background-color: #0052a3;
    }
    .metric-card {
        background-color: #f0f2f6;
        border-radius: 10px;
        padding: 1rem;
        box-shadow: 2px 2px 5px rgba(0,0,0,0.1);
    }
    .st-emotion-cache-1kyxreq {
        margin-top: -75px;
    }
    </style>
""", unsafe_allow_html=True)


#puse un agente adaptativo para no recargar
class AdaptiveAgent:
    def __init__(self):
        self.learning_rate = 0.1
        self.performance_history = []
        self.adaptation_threshold = 0.7
        self.state = {}
        self.load_state()
    
    def load_state(self):
        try:
            if os.path.exists('agent_state.json'):
                with open('agent_state.json', 'r') as f:
                    self.state = json.load(f)
        except Exception as e:
            st.warning(f"No se pudo cargar el estado del agente: {e}")
    
    def save_state(self):
        try:
            with open('agent_state.json', 'w') as f:
                json.dump(self.state, f)
        except Exception as e:
            st.warning(f"No se pudo guardar el estado del agente: {e}")
    
    def update_parameters(self, performance_metrics):
        """Actualiza los parametros basado en el rendimiento"""
        current_performance = np.mean(performance_metrics)
        self.performance_history.append(current_performance)
        
        if len(self.performance_history) > 1:
            performance_trend = self.performance_history[-1] - self.performance_history[-2]
            
            if performance_trend < 0:
                self.learning_rate *= 0.9  # Reduce learning rate if performance decreases
            else:
                self.learning_rate = min(self.learning_rate * 1.1, 0.5)
        
        self.state['learning_rate'] = self.learning_rate
        self.state['last_performance'] = current_performance
        self.save_state()
        
        return {
            'learning_rate': self.learning_rate,
            'performance_trend': performance_trend if len(self.performance_history) > 1 else 0
        }

    def suggest_parameters(self, current_params):
        """Sugiere ajustes a los parametros basado en el aprendizaje"""
        if not self.performance_history:
            return current_params
        
        recent_performance = np.mean(self.performance_history[-5:]) if len(self.performance_history) >= 5 else np.mean(self.performance_history)
        
        if recent_performance < self.adaptation_threshold:
            suggested_params = current_params.copy()
            suggested_params['min_alumnos'] = max(5, current_params['min_alumnos'] - 1)
            suggested_params['max_carga_profesor'] = min(25, current_params['max_carga_profesor'] + 1)
            return suggested_params
        
        return current_params

class ScheduleOptimizer:
    def __init__(self):
        self.scaler = StandardScaler()
        self.horarios_generados = []
        self.success_rate = 0.0  # Añadido
        self.best_model = None
        self.is_fitted = False
        self.best_model_score = 0.0
        self.last_training_date = None
        self.feature_names = [
            'experiencia', 'calificacion_alumno', 'alumnos', 
            'bloques', 'horarios_disponibles', 'capacidad_salon',
            'conflictos_horario', 'carga_profesor'
        ]
        self.load_model()
        self.adaptive_agent = AdaptiveAgent()
        self.performance_history = []
        self.pattern_database = {}
        self.slot_duration = 45
        self.model_params = {'test_size': 0.2}
        
        self.grupos_profesor = {} 
        self.grupo_counter = {} 
        self.grupos_asignados = {}  
        self.grupos_por_materia = {}  
        # Modificar la definición de jornadas para ser más flexible
        self.JORNADAS = {
            'mañana': (self.parse_time("06:00"), self.parse_time("12:59")),
            'tarde': (self.parse_time("13:00"), self.parse_time("17:59")),
            'noche': (self.parse_time("18:00"), self.parse_time("22:00"))
        }
        
        # Modificar las horas permitidas para incluir transiciones
        self.HORAS_PERMITIDAS = {
            'mañana': ["06:00", "07:30", "09:00", "10:30", "12:00"],
            'tarde': ["13:30", "15:00", "16:30" ],
            'noche': ["18:30", "19:15", "20:00", "20:45", "21:30"]
        }
        
        self.JORNADA_PREFIXES = {
            'mañana': 'B',
            'tarde': 'C',
            'noche': 'E'
        }

        self.grupo_profesor_map = {}  
        self.profesor_grupos = {}     
        self.ultimo_grupo_materia = {} 
        self.horarios_generados = []
        self.load_horarios_history()
        self.materia_grupos = {}  
        self.materia_base_codes = {}
        self.hora_usage = {}
        self.materia_codigos = {}
        
    def _get_jornadas_for_timerange(self, hora_inicio, hora_fin):
        """
        Determina qué jornadas están incluidas en un rango de tiempo dado
        
        Args:
            hora_inicio: objeto time o string con la hora de inicio
            hora_fin: objeto time o string con la hora de fin
        
        Returns:
            list: Lista de jornadas que aplican
        """
        if isinstance(hora_inicio, str):
            hora_inicio = self.parse_time(hora_inicio)
        if isinstance(hora_fin, str):
            hora_fin = self.parse_time(hora_fin)
        
        jornadas_aplicables = []
        for jornada, (inicio_jornada, fin_jornada) in self.JORNADAS.items():
            # Verificar si hay solapamiento entre el rango dado y la jornada
            if (hora_inicio <= fin_jornada and hora_fin >= inicio_jornada):
                jornadas_aplicables.append(jornada)
                
        return jornadas_aplicables
    def _get_horas_disponibles(self, hora_inicio, hora_fin, intervalo=45):
        """
        Genera slots de tiempo disponibles entre hora_inicio y hora_fin
        
        """
        if isinstance(hora_inicio, str):
            hora_inicio = self.parse_time(hora_inicio)
        if isinstance(hora_fin, str):
            hora_fin = self.parse_time(hora_fin)
        
        horas_disponibles = []
        hora_actual = hora_inicio
        
        while hora_actual < hora_fin:
            # Convertir a datetime para poder sumar minutos
            dt_actual = datetime.combine(datetime.min, hora_actual)
            # Verificar si aún hay espacio para un slot completo
            if (datetime.combine(datetime.min, hora_fin) - dt_actual).seconds/60 >= intervalo:
                horas_disponibles.append(hora_actual.strftime("%H:%M"))
            # Avanzar al siguiente slot
            dt_actual += timedelta(minutes=intervalo)
            hora_actual = dt_actual.time()
        
        return sorted(horas_disponibles)

    def save_configuration(self, config):
        """Guarda la configuración del sistema"""
        try:
            if not os.path.exists('config'):
                os.makedirs('config')
            
            # Convertir objetos time a string antes de guardar
            clean_config = {}
            for section, params in config.items():
                clean_config[section] = {}
                for key, value in params.items():
                    if isinstance(value, time):
                        clean_config[section][key] = value.strftime("%H:%M")
                    elif isinstance(value, (list, set)):
                        clean_config[section][key] = list(value)
                    else:
                        clean_config[section][key] = value
            
            # Guardar en archivo
            with open('config/system_config.json', 'w', encoding='utf-8') as f:
                json.dump(clean_config, f, indent=4, default=str)
            
            # Actualizar las propiedades del optimizador
            self._update_optimizer_properties(config)
            
            return True, "Configuración guardada exitosamente"
        except Exception as e:
            return False, f"Error al guardar la configuración: {str(e)}"

    def _prepare_config_for_save(self, config):
        """Prepara la configuración para ser guardada en JSON"""
        clean_config = {}
        for section, params in config.items():
            clean_config[section] = {}
            for key, value in params.items():
                # Convertir time objects a string
                if isinstance(value, time):
                    clean_config[section][key] = value.strftime("%H:%M")
                # Convertir sets a list
                elif isinstance(value, set):
                    clean_config[section][key] = list(value)
                else:
                    clean_config[section][key] = value
        return clean_config

    def _update_optimizer_properties(self, config):
        """Actualiza las propiedades del optimizador con la nueva configuración"""
       
        if 'basic' in config:
            self.slot_duration = config['basic'].get('slot_duration', self.slot_duration)
            self.min_alumnos = config['basic'].get('min_alumnos', 10)
            self.max_carga_profesor = config['basic'].get('max_carga_profesor', 20)

        # Actualizar parametros del agente adaptativo
        if 'adaptive' in config and hasattr(self, 'adaptive_agent'):
            self.adaptive_agent.learning_rate = config['adaptive'].get('learning_rate', 
                                                                     self.adaptive_agent.learning_rate)
            self.adaptive_agent.adaptation_threshold = config['adaptive'].get('adaptation_threshold',
                                                                           self.adaptive_agent.adaptation_threshold)

    def load_configuration(self):
        """Carga la configuración guardada"""
        try:
            if os.path.exists('config/system_config.json'):
                with open('config/system_config.json', 'r', encoding='utf-8') as f:
                    return json.load(f)
            return None
        except Exception as e:
            st.error(f"Error al cargar la configuración: {str(e)}")
            return None

    def save_horarios_history(self):
        """Guarda el historial de horarios generados"""
        try:
            if not os.path.exists('models'):
                os.makedirs('models')
            with open('models/horarios_history.json', 'w') as f:
                # Asegurarse de que los datos son serializables
                history_data = []
                for item in self.horarios_generados:
                    # Crear una copia limpia del item que sea serializable
                    clean_item = {
                        'fecha': item['fecha'],
                        'status': item['status'],
                        'num_clases': item['num_clases'],
                        'warnings': item['warnings'],
                        'optimization_params': {
                            k: str(v) if isinstance(v, (list, dict)) else v 
                            for k, v in item['optimization_params'].items()
                        }
                    }
                    history_data.append(clean_item)
                json.dump(history_data, f, indent=2)
        except Exception as e:
            st.warning(f"No se pudo guardar el historial de horarios: {e}")

    def load_horarios_history(self):
        """Carga el historial de horarios generados"""
        try:
            if os.path.exists('models/horarios_history.json'):
                with open('models/horarios_history.json', 'r') as f:
                    self.horarios_generados = json.load(f)
            else:
                self.horarios_generados = []
        except Exception as e:
            st.warning(f"No se pudo cargar el historial de horarios: {e}")
            self.horarios_generados = []
                
    def save_model(self):
        """Guarda el modelo y sus métricas"""
        if self.is_fitted:
            try:
                if not os.path.exists('models'):
                    os.makedirs('models')
                
                # Guardar modelo y scaler
                joblib.dump(self.best_model, 'models/best_model.joblib')
                joblib.dump(self.scaler, 'models/scaler.joblib')
                joblib.dump(self.is_fitted, 'models/is_fitted.joblib')
                
                # Calcular tasa de éxito actualizada
                if self.horarios_generados:
                    optimal_count = len([h for h in self.horarios_generados 
                                    if h['status'] == 'OPTIMAL'])
                    self.success_rate = optimal_count / len(self.horarios_generados)
                
                # Guardar métricas
                metrics = {
                    'feature_names': self.feature_names,
                    'performance_history': self.performance_history,
                    'best_model_score': float(self.best_model_score),
                    'last_training_date': self.last_training_date,
                    'success_rate': float(self.success_rate),
                    'last_update': datetime.now().isoformat()
                }
                
                with open('models/metrics.json', 'w') as f:
                    json.dump(metrics, f, default=str)
                
                return True
            except Exception as e:
                st.error(f"Error al guardar el modelo: {str(e)}")
                return False
        return False
                

    def load_model(self):
        """Carga el modelo y sus métricas"""
        try:
            if os.path.exists('models/best_model.joblib') and \
            os.path.exists('models/scaler.joblib') and \
            os.path.exists('models/is_fitted.joblib'):
                
                self.best_model = joblib.load('models/best_model.joblib')
                self.scaler = joblib.load('models/scaler.joblib')
                self.is_fitted = joblib.load('models/is_fitted.joblib')
                
                # Cargar métricas adicionales
                if os.path.exists('models/metrics.json'):
                    with open('models/metrics.json', 'r') as f:
                        metrics = json.load(f)
                        self.best_model_score = float(metrics.get('best_model_score', 0.0))
                        self.last_training_date = metrics.get('last_training_date', None)
                        self.performance_history = metrics.get('performance_history', [])
                        
                # Cargar historial de horarios
                if os.path.exists('models/horarios_history.json'):
                    with open('models/horarios_history.json', 'r') as f:
                        self.horarios_generados = json.load(f)
                        # Calcular tasa de éxito
                        if self.horarios_generados:
                            optimal_count = len([h for h in self.horarios_generados 
                                            if h['status'] == 'OPTIMAL'])
                            self.success_rate = optimal_count / len(self.horarios_generados)
                        
                return True
            return False
        except Exception as e:
            st.error(f"Error al cargar el modelo: {str(e)}")
            return False
        

    
    @st.cache_data
    def get_data(_self, endpoint):  # odtengo los datos usanto los get, esto lo guardo en cache
        try:
            response = requests.get(f"{BASE_URL}/{endpoint}")
            response.raise_for_status()
            return response.json()
        except requests.RequestException as e:
            st.error(f"Error al obtener datos de {endpoint}: {str(e)}")
            return None
        
    #preparo los datos

    def prepare_features(self, df_profesores, df_materias, df_salones, df_horarios, df_profesor_materia):
        features = []
        labels = []
        conflicts = []

        # Cargo el dicionario para el modelo por medio del id del profesor
        carga_profesor = df_profesor_materia.groupby('profesor_id').size().to_dict()
        
        for _, prof_mat in df_profesor_materia.iterrows():
            profesor = df_profesores[df_profesores['id'] == prof_mat['profesor_id']].iloc[0]
            materia = df_materias[df_materias['id'] == prof_mat['materia_id']].iloc[0]
            
            # calculo si hay conflictos de ids entre los datos
            horarios_prof = df_horarios[df_horarios['profesor_id'] == prof_mat['profesor_id']]
            conflictos = self.calcular_conflictos(horarios_prof)
            
            # Para cada salón disponible
            for _, salon in df_salones.iterrows():
                if salon['capacidad_alumnos'] >= materia['alumnos']:
                    feature = [
                        prof_mat['experiencia'],
                        prof_mat['calificacion_alumno'],
                        materia['alumnos'],
                        materia['bloques'],
                        len(horarios_prof),
                        salon['capacidad_alumnos'],
                        conflictos,
                        carga_profesor.get(prof_mat['profesor_id'], 0)
                    ]
                    
                    features.append(feature)
                    labels.append(1)  # Combinación válida
                    conflicts.append(conflictos)

        # Generar ejemplos negativos más realistas
        negative_examples = self.generate_negative_examples(
            df_profesores, df_materias, df_salones, df_horarios, 
            df_profesor_materia, len(features)
        )
        
        features.extend(negative_examples[0])
        labels.extend(negative_examples[1])
        conflicts.extend(negative_examples[2])

        return np.array(features), np.array(labels), conflicts
    
    def _get_siguiente_codigo_grupo(self, materia_id, profesor_id, jornada):
        """
        Obtiene el siguiente código de grupo disponible para una materia y profesor
        """
        # Inicializar el tracking para esta materia si no existe
        if materia_id not in self.grupos_por_materia:
            self.grupos_por_materia[materia_id] = {}
            
        # Inicializar el contador para este profesor si no existe
        if profesor_id not in self.grupos_por_materia[materia_id]:
            self.grupos_por_materia[materia_id][profesor_id] = {
                'mañana': [],
                'tarde': [],
                'noche': []
            }
        
        grupos_profesor = self.grupos_por_materia[materia_id][profesor_id][jornada]
        
        # Si no hay grupos para este profesor en esta jornada, empezar con 100
        if not grupos_profesor:
            nuevo_numero = 100
        else:
            # Encontrar el siguiente número disponible
            numeros_usados = set(int(grupo[1:]) for grupo in grupos_profesor)
            nuevo_numero = max(numeros_usados) + 1 if numeros_usados else 100
        
        # Crear el nuevo código
        prefix = self.JORNADA_PREFIXES[jornada]
        nuevo_codigo = f"{prefix}{nuevo_numero}"
        
        # Registrar el nuevo código
        grupos_profesor.append(nuevo_codigo)
        
        return nuevo_codigo
    
    def _get_horas_jornada(self, jornada):
        """Obtiene las horas permitidas para una jornada específica"""
        inicio, fin = self.JORNADAS[jornada]
        horas_jornada = [
            hora for hora in self.HORAS_PERMITIDAS 
            if inicio <= self.parse_time(hora) <= fin
        ]
        
        # Debug info
        print(f"Horas disponibles para jornada {jornada}: {horas_jornada}")
        return horas_jornada
    
    def _asignar_grupo_profesor(self, materia_id, profesor_id, jornada):
        """
        Asigna o recupera un grupo para un profesor y materia específicos,
        manteniendo un contador por profesor y materia
        """
        # Clave única para la materia y profesor
        clave = f"{materia_id}_{profesor_id}"
        
        # Inicializar el contador de grupos para esta combinación si no existe
        if clave not in self.profesor_grupos:
            self.profesor_grupos[clave] = {
                'mañana': 0,
                'tarde': 0,
                'noche': 0
            }
        
        # Incrementar el contador para la jornada específica
        self.profesor_grupos[clave][jornada] += 1
        
        # Generar el código del grupo
        prefijo = self.JORNADA_PREFIXES[jornada]
        numero = 90 + self.profesor_grupos[clave][jornada]  # Empezamos desde 91
        
        return f"{prefijo}{numero}"


    def _generar_codigo_grupo(self, materia_id, jornada):
        """
        Genera un nuevo código de grupo único para una materia
        """
        if materia_id not in self.ultimo_grupo_materia:
            self.ultimo_grupo_materia[materia_id] = 90  # Empezamos en 91
            
        self.ultimo_grupo_materia[materia_id] += 1
        numero = self.ultimo_grupo_materia[materia_id]
        
        # Prefijo según la jornada (B: mañana, C: tarde, E: noche)
        prefix = self.JORNADA_PREFIXES[jornada]
        
        return f"{prefix}{numero}"
        
    def _get_jornada_from_hora(self, hora):
        """Determina la jornada basada en la hora"""
        hora_obj = self.parse_time(hora) if isinstance(hora, str) else hora
        for jornada, (inicio, fin) in self.JORNADAS.items():
            if inicio <= hora_obj <= fin:
                return jornada
        return None

    def _get_next_available_hora(self, jornada, dia):
        """Obtiene una hora disponible aleatoria en la jornada especificada"""
        if dia not in self.hora_usage:
            self.hora_usage[dia] = {hora: 0 for hora in self.HORAS_PERMITIDAS}
        
        horas_jornada = []
        # Filtrar horas según la jornada
        if jornada == 'mañana':
            horas_jornada = self.HORAS_PERMITIDAS['mañana']
        elif jornada == 'tarde':
            horas_jornada = self.HORAS_PERMITIDAS['tarde']
        elif jornada == 'noche':
            horas_jornada = self.HORAS_PERMITIDAS['noche']
        
        if not horas_jornada:
            return None
            
        # Encontrar las horas menos utilizadas
        min_usage = min(self.hora_usage[dia][hora] for hora in horas_jornada)
        horas_disponibles = [
            hora for hora in horas_jornada 
            if self.hora_usage[dia][hora] == min_usage
        ]
        
        # Seleccionar una hora aleatoria entre las menos utilizadas
        if horas_disponibles:
            hora_seleccionada = random.choice(horas_disponibles)
            self.hora_usage[dia][hora_seleccionada] += 1
            return hora_seleccionada
                
        return random.choice(horas_jornada) 

    def hay_solapamiento(self, inicio1, fin1, inicio2, fin2):
        """Verifica si hay solapamiento entre dos rangos de tiempo"""
        if isinstance(inicio1, str):
            inicio1 = self.parse_time(inicio1)
        if isinstance(fin1, str):
            fin1 = self.parse_time(fin1)
        if isinstance(inicio2, str):
            inicio2 = self.parse_time(inicio2)
        if isinstance(fin2, str):
            fin2 = self.parse_time(fin2)

        return (inicio1 < fin2 and fin1 > inicio2)
    
    def calcular_conflictos(self, horarios_prof):
        if horarios_prof.empty:
            return 0
        
        conflictos = 0
        horarios_list = horarios_prof.values.tolist()
        
        for i in range(len(horarios_list)):
            for j in range(i + 1, len(horarios_list)):
                if self.hay_solapamiento(
                    horarios_list[i][2], horarios_list[i][3],  # hora_inicio, hora_fin del primer horario
                    horarios_list[j][2], horarios_list[j][3]   # hora_inicio, hora_fin del segundo horario
                ):
                    conflictos += 1
        
        return conflictos

    def generate_negative_examples(self, df_profesores, df_materias, df_salones, 
                                 df_horarios, df_profesor_materia, num_samples):
        features = []
        labels = []
        conflicts = []
        
        for _ in range(num_samples):
            profesor = df_profesores.sample(1).iloc[0]
            materia = df_materias.sample(1).iloc[0]
            salon = df_salones.sample(1).iloc[0]
            
            # Verificar si es una combinación inválida
            if df_profesor_materia[(df_profesor_materia['profesor_id'] == profesor['id']) & 
                                 (df_profesor_materia['materia_id'] == materia['id'])].empty:
                
                horarios_prof = df_horarios[df_horarios['profesor_id'] == profesor['id']]
                conflictos = self.calcular_conflictos(horarios_prof)
                
                feature = [
                    np.random.randint(1, 5),  # experiencia aleatoria, no se es mas random
                    np.random.randint(1, 5),  # calificación del teacher
                    materia['alumnos'],
                    materia['bloques'],
                    len(horarios_prof),
                    salon['capacidad_alumnos'],
                    conflictos,
                    len(df_profesor_materia[df_profesor_materia['profesor_id'] == profesor['id']])
                ]
                
                features.append(feature)
                labels.append(0)
                conflicts.append(conflictos)
        
        return features, labels, conflicts

    def train_model(self, X, y, model_params):
        """
        Entrena el modelo usando los parámetros especificados y retorna los resultados

        Args:
            X (np.array): Features de entrenamiento
            y (np.array): Labels de entrenamiento
            model_params (dict): Parametros del modelo y entrenamiento

        Returns:
            dict: Resultados del entrenamiento incluyendo metricas y parametros óptimos
        """
        try:
            # Escalar los datos
            X_scaled = self.scaler.fit_transform(X)
            
            # Split de datos
            X_train, X_test, y_train, y_test = train_test_split(
                X_scaled, y, 
                test_size=model_params.get('test_size', 0.2),
                random_state=model_params.get('random_state', 42)
            )
            
            # Configurar el modelo segun el tipo
            if model_params.get('model_type', '').lower() == 'knn':
                base_model = KNeighborsClassifier()
                param_grid = {
                    'n_neighbors': [model_params.get('n_neighbors', 5)],
                    'weights': [model_params.get('weights', 'uniform')],
                    'metric': [model_params.get('metric', 'euclidean')]
                }
            else:  # Random Forest por defecto
                base_model = RandomForestClassifier(
                    random_state=model_params.get('random_state', 42)
                )
                param_grid = {
                    'n_estimators': [model_params.get('n_estimators', 100)],
                    'max_depth': [model_params.get('max_depth', 10)],
                    'min_samples_split': [model_params.get('min_samples_split', 2)],
                    'min_samples_leaf': [model_params.get('min_samples_leaf', 1)]
                }
            
            # Configurar y ejecutar GridSearchCV
            grid_search = GridSearchCV(
                base_model, 
                param_grid, 
                cv=model_params.get('cv_folds', 5),
                scoring='accuracy',  # Cambiado a accuracy para mejor interpretabilidad, esto se puede modificar
                n_jobs=-1,
                verbose=1
            )
            
            # Entrenar el modelo
            grid_search.fit(X_train, y_train)
            
            # Guardar el mejor modelo
            self.best_model = grid_search.best_estimator_
            self.is_fitted = True
            
            # Realizar predicciones en el conjunto de prueba
            y_pred = self.best_model.predict(X_test)
            y_pred_train = self.best_model.predict(X_train)
            
            # Calcular métricas detalladas
            train_accuracy = np.mean(y_pred_train == y_train)
            test_accuracy = np.mean(y_pred == y_test)
            
            # Actualizar las métricas del modelo
            self.best_model_score = test_accuracy
            self.last_training_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Calcular matriz de confusión y reporte de clasificación
            conf_matrix = confusion_matrix(y_test, y_pred)
            class_report = classification_report(y_test, y_pred, output_dict=True)
            
            # Calcular métricas adicionales
            feature_importance = {}
            if isinstance(self.best_model, RandomForestClassifier):
                feature_importance = dict(zip(
                    self.feature_names,
                    self.best_model.feature_importances_
                ))
            
            # Crear resumen de resultados
            results = {
                'best_params': grid_search.best_params_,
                'best_score': self.best_model_score,
                'train_accuracy': train_accuracy,
                'test_accuracy': test_accuracy,
                'confusion_matrix': conf_matrix,
                'classification_report': class_report,
                'training_date': self.last_training_date,
                'feature_importance': feature_importance,
                'model_type': model_params.get('model_type', 'random_forest'),
                'training_metrics': {
                    'train_size': len(X_train),
                    'test_size': len(X_test),
                    'cv_folds': model_params.get('cv_folds', 5),
                    'grid_scores': grid_search.cv_results_['mean_test_score'].tolist()
                }
            }
            
            # Actualizar historial de rendimiento
            self.performance_history.append({
                'date': self.last_training_date,
                'score': self.best_model_score,
                'train_accuracy': train_accuracy,
                'test_accuracy': test_accuracy,
                'model_type': model_params.get('model_type', 'random_forest'),
                'params': grid_search.best_params_
            })
            
            # Guardar el modelo y las métricas
            if self.save_model():
                st.success("✅ Modelo guardado exitosamente")
                
                # Mostrar métricas de entrenamiento
                st.subheader("📊 Métricas de Entrenamiento")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric(
                        label="Precisión en Entrenamiento",
                        value=f"{train_accuracy:.2%}",
                        help="Precisión del modelo en los datos de entrenamiento"
                    )
                with col2:
                    st.metric(
                        label="Precisión en Prueba",
                        value=f"{test_accuracy:.2%}",
                        help="Precisión del modelo en los datos de prueba"
                    )
                with col3:
                    st.metric(
                        label="Mejor Score CV",
                        value=f"{grid_search.best_score_:.2%}",
                        help="Mejor score obtenido en la validación cruzada"
                    )
                
                # Mostrar gráficos de evolución del entrenamiento
                if len(results['training_metrics']['grid_scores']) > 1:
                    fig_scores = px.line(
                        x=range(1, len(results['training_metrics']['grid_scores']) + 1),
                        y=results['training_metrics']['grid_scores'],
                        title='Evolución de Scores durante Grid Search',
                        labels={'x': 'Iteración', 'y': 'Score'}
                    )
                    st.plotly_chart(fig_scores)
                
                # Mostrar importancia de características si está disponible
                if feature_importance:
                    fig_importance = px.bar(
                        x=list(feature_importance.keys()),
                        y=list(feature_importance.values()),
                        title='Importancia de Características',
                        labels={'x': 'Características', 'y': 'Importancia'}
                    )
                    st.plotly_chart(fig_importance)
            
            return results

        except Exception as e:
            st.error(f"Error durante el entrenamiento: {str(e)}")
            raise e

    def generate_schedule(self, df_profesores, df_materias, df_salones, df_horarios, df_profesor_materia, optimization_params):
        """
        Genera un horario optimizado basado en los parámetros y restricciones guardadas

        """
        self.grupo_profesor_map = {}
        self.profesor_grupos = {}
        self.ultimo_grupo_materia = {}

        

        try:
            # Cargar configuración guardada
            saved_config = self.load_configuration()
            
            # Inicializar parámetros base
            base_params = {
                'slot_duration': 45,
                'min_alumnos': 10,
                'max_carga_profesor': 20,
                'dias_habiles': ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sabado"],
                'horario_inicio': "06:00",
                'horario_fin': "23:45",
                'optimization_level': "Medio",
                'max_iterations': 1000,
                'allow_overlap': False,
                'priority_rules': ["Materias con más alumnos"],
                'max_clases_consecutivas': 3,
                'min_descanso': 15,
                'max_ventanas': 2,
                'distancia_maxima': 100,
            }

            # Actualizar con configuración guardada si existe
            if saved_config:
                if 'basic' in saved_config:
                    base_params.update(saved_config['basic'])
                if 'advanced' in saved_config:
                    base_params.update(saved_config['advanced'])
                if 'restrictions' in saved_config:
                    base_params.update(saved_config['restrictions'])

            # Actualizar con parámetros recibidos (tienen prioridad)
            base_params.update(optimization_params)

            # Verificar si el modelo está entrenado
            if not self.is_fitted:
                return {
                    "status": "ERROR",
                    "horario_generado": [],
                    "warnings": [],
                    "errors": ["El modelo no ha sido entrenado. Por favor, entrene el modelo primero."]
                }

            horario_generado = []
            warnings = []
            errors = []
            
            # Ordenar materias según la prioridad
            df_materias_sorted = df_materias.sort_values(
                ['alumnos', 'bloques'], 
                ascending=[False, False]
            )
            
            # Iniciar generación de horario
            for _, materia in df_materias_sorted.iterrows():
                if materia['alumnos'] < base_params['min_alumnos']:
                    warnings.append(f"Materia {materia['nombre']} no tiene suficientes alumnos")
                    continue
                    
                clases_asignadas = self.asignar_clases(
                    materia, df_profesores, df_salones, df_horarios,
                    df_profesor_materia, base_params, horario_generado
                )
                
                if clases_asignadas < materia['bloques']:
                    warnings.append(
                        f"No se pudieron asignar todos los horarios para {materia['nombre']}"
                    )

            # Determinar el estado del resultado
            status = "OPTIMAL" if len(warnings) == 0 else "FEASIBLE"
            
            resultado = {
                "status": status,
                "horario_generado": horario_generado,
                "warnings": warnings,
                "errors": errors,
                "params_used": base_params
            }

            # Registrar el horario generado
            self.horarios_generados.append({
                'fecha': datetime.now().isoformat(),
                'status': status,
                'num_clases': len(horario_generado),
                'warnings': len(warnings),
                'optimization_params': {k: str(v) if isinstance(v, (list, dict)) else v 
                                    for k, v in base_params.items()}
            })
            self.save_horarios_history()
            
            return resultado

        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            return {
                "status": "ERROR",
                "horario_generado": [],
                "warnings": [],
                "errors": [f"Error en la generación del horario: {str(e)}", f"Detalle: {error_detail}"]
            }

    def _filter_horarios_by_config(self, df_horarios, config):
        # Convertir tiempo una sola vez
        hora_inicio = self.parse_time(config['horario_inicio'])
        hora_fin = self.parse_time(config['horario_fin'])
        # Usar vectorización para mejor rendimiento
        return df_horarios[
            df_horarios['dia'].isin(config['dias_habiles']) &
            df_horarios['hora_inicio'].apply(self.parse_time).between(hora_inicio, hora_fin)
        ]

    def _sort_materias_by_priority(self, df_materias, priority_rules):
        """Ordena materias según reglas de prioridad"""
        sort_columns = []
        ascending = []
        
        if "Materias con más alumnos" in priority_rules:
            sort_columns.extend(['alumnos', 'bloques'])
            ascending.extend([False, False])
        if "Profesores más experimentados" in priority_rules:
            sort_columns.append('experiencia_requerida')
            ascending.append(False)
            
        return df_materias.sort_values(sort_columns, ascending=ascending)

    def _verify_additional_restrictions(self, horario_generado, config):
        # Usar diccionarios para O(1) 
        profesor_horarios = {}
        warnings = []
        
        # Agrupar por profesor una sola vez
        for clase in horario_generado:
            prof_id = clase['profesor_id']
            profesor_horarios.setdefault(prof_id, []).append(clase)
        
        for prof_id, clases in profesor_horarios.items():
            clases_ordenadas = sorted(clases, key=lambda x: (x['dia_semana'], x['hora_inicio']))
            
            # Validación por ventanas en O(n)
            ventanas = self._contar_ventanas(clases_ordenadas)
            if ventanas > config['max_ventanas']:
                warnings.append(f"Profesor {prof_id} excede máximo de ventanas")
                
        return warnings

    def _is_consecutive(self, clase1, clase2, min_descanso=15):
        """Verifica si dos clases son consecutivas"""
        if clase1['dia_semana'] != clase2['dia_semana']:
            return False
            
        fin_clase1 = datetime.strptime(clase1['hora_fin'], '%H:%M').time()
        inicio_clase2 = datetime.strptime(clase2['hora_inicio'], '%H:%M').time()
        
        diferencia = datetime.combine(datetime.today(), inicio_clase2) - \
                        datetime.combine(datetime.today(), fin_clase1)
                        
        return diferencia.total_seconds() / 60 < min_descanso

    def _is_window(self, clase1, clase2):
        """Verifica si hay una ventana entre clases"""
        if clase1['dia_semana'] != clase2['dia_semana']:
            return False
            
        fin_clase1 = datetime.strptime(clase1['hora_fin'], '%H:%M').time()
        inicio_clase2 = datetime.strptime(clase2['hora_inicio'], '%H:%M').time()
        
        diferencia = datetime.combine(datetime.today(), inicio_clase2) - \
                        datetime.combine(datetime.today(), fin_clase1)
                        
        return diferencia.total_seconds() / 60 > 90  # Ventana si hay más de 90 min

    def _dividir_grupos(self, materia, df_salones, params):
        """
        Divide una materia en grupos asegurando que cada grupo tenga un único profesor
        y se mantenga en una única jornada.
        """
        alumnos_totales = materia['alumnos']
        grupos = []
        alumnos_restantes = alumnos_totales
        grupo_counter = 1  # Contador simple para el número de grupo
        
        # Ordenar salones por capacidad (de mayor a menor)
        salones_ordenados = df_salones.sort_values('capacidad_alumnos', ascending=False)
        
        while alumnos_restantes > 0 and not salones_ordenados.empty:
            for _, salon in salones_ordenados.iterrows():
                # Determinar tamaño del grupo basado en la capacidad del salón
                tam_grupo = min(salon['capacidad_alumnos'], alumnos_restantes)
                
                if tam_grupo >= params['min_alumnos']:
                    grupo = {
                        'alumnos': tam_grupo,
                        'salon_id': salon['id'],
                        'profesor_id': None,  # Se asignará durante la generación de horarios
                        'jornada': None,      # Se asignará durante la generación de horarios
                        'grupo_numero': grupo_counter  # Asignar número de grupo
                    }
                    grupos.append(grupo)
                    alumnos_restantes -= tam_grupo
                    grupo_counter += 1
                    break
            
            if alumnos_restantes < params['min_alumnos'] and grupos:
                # Agregar remanente al último grupo
                grupos[-1]['alumnos'] += alumnos_restantes
                alumnos_restantes = 0
                
        return grupos

    def asignar_clases(self, materia, df_profesores, df_salones, df_horarios, df_profesor_materia, params, horario_generado):
        """
        Asigna clases respetando las preferencias de jornada y restricciones del sistema.
        
     
        """
        grupos = self._dividir_grupos(materia, df_salones, params)
        if not grupos:
            return 0
                
        clases_asignadas = 0
        horarios_ocupados = self._crear_indice_horarios_ocupados(horario_generado)
        
        # Obtener todos los profesores elegibles una vez
        profesores_elegibles = self._obtener_profesores_elegibles(
            materia, df_profesores, df_profesor_materia, horario_generado, params
        )
        
        if not profesores_elegibles:
            return 0
        
        # Obtener las jornadas preferidas del usuario
        jornadas_preferidas = [j.lower() for j in params.get('preferencia_horario', ['mañana', 'tarde', 'noche'])]
        if not jornadas_preferidas:  # Si no hay preferencias, usar todas
            jornadas_preferidas = ['mañana', 'tarde']
        
        # Procesar cada grupo
        for grupo in grupos:
            # Seleccionar el profesor con menor carga
            profesor = min(
                profesores_elegibles,
                key=lambda p: self.get_carga_actual(p['id'], horario_generado)
            )
                
            grupo['profesor_id'] = profesor['id']
                
            # Intentar asignar en las jornadas preferidas
            for jornada in jornadas_preferidas:
                bloques_asignados = 0
                horarios_tentativos = []
                    
                # Filtrar horarios del profesor para la jornada
                horarios_prof = df_horarios[df_horarios['profesor_id'] == profesor['id']]
                horarios_jornada = self._filtrar_horarios_por_jornada(horarios_prof, jornada)
                    
                if horarios_jornada.empty:
                    continue
                    
                # Procesar cada día disponible
                dias_disponibles = sorted(set(params['dias_habiles']))
                bloques_por_asignar = materia['bloques']
                    
                for dia in dias_disponibles:
                    if bloques_por_asignar <= 0:
                        break
                            
                    # Obtener horas disponibles para este día
                    horas_disponibles = sorted(set(
                        horarios_jornada[horarios_jornada['dia'] == dia]['hora_inicio']
                    ))
                        
                    for hora in horas_disponibles:
                        if bloques_por_asignar <= 0:
                            break
                                
                        hora_inicio = self.parse_time(hora)
                        hora_fin = (datetime.combine(datetime.today(), hora_inicio) + 
                                timedelta(minutes=90)).time()
                            
                        # Verificar disponibilidad y restricciones
                        if ((dia, hora, hora_fin.strftime("%H:%M")) not in horarios_ocupados and
                            self._cumple_restricciones_tiempo(
                                profesor['id'], dia, hora_inicio, hora_fin,
                                horario_generado + horarios_tentativos, params
                            )):
                            # Generar código de grupo
                            prefix = self.JORNADA_PREFIXES[jornada]
                            grupo_codigo = f"{prefix}{materia['id']:03d}-{str(grupo['grupo_numero']).zfill(2)}"
                                
                            # Crear bloques de clase (2 bloques de 45 minutos)
                            nuevas_clases = []
                            for i in range(2):
                                hora_clase_inicio = (datetime.combine(datetime.today(), hora_inicio) + 
                                                timedelta(minutes=45*i)).time()
                                hora_clase_fin = (datetime.combine(datetime.today(), hora_clase_inicio) + 
                                                timedelta(minutes=45)).time()
                                    
                                nueva_clase = {
                                    'grupo': grupo_codigo,
                                    'dia_semana': dia,
                                    'hora_inicio': hora_clase_inicio.strftime('%H:%M'),
                                    'hora_fin': hora_clase_fin.strftime('%H:%M'),
                                    'alumnos': grupo['alumnos'],
                                    'materia_id': materia['id'],
                                    'salon_id': grupo['salon_id'],
                                    'profesor_id': profesor['id']
                                }
                                nuevas_clases.append(nueva_clase)
                                
                            horarios_tentativos.extend(nuevas_clases)
                            bloques_por_asignar -= 1
                            bloques_asignados += 1
                    
                # Si se asignaron todos los bloques necesarios
                if bloques_por_asignar <= 0:
                    grupo['jornada'] = jornada
                    horario_generado.extend(horarios_tentativos)
                    clases_asignadas += bloques_asignados
                    break
        
        return clases_asignadas

    def _es_horario_disponible(self, profesor_id, hora, horario_generado, params):
        """
        Verifica si un horario está disponible para un profesor
        """
        hora_obj = self.parse_time(hora)
        hora_fin = (datetime.combine(datetime.min, hora_obj) + 
                   timedelta(minutes=90)).time()
                   
        # Verificar solapamientos
        for clase in horario_generado:
            if (clase['profesor_id'] == profesor_id and
                clase['dia_semana'] == params.get('dia', 'Lunes') and
                self.hay_solapamiento(
                    self.parse_time(clase['hora_inicio']),
                    self.parse_time(clase['hora_fin']),
                    hora_obj,
                    hora_fin
                )):
                return False
        
        return True

    def _obtener_jornada(self, hora):
        """Determina la jornada según la hora"""
        if 6 <= hora.hour < 13:
            return 'Mañana'
        elif 13 <= hora.hour < 18:
            return 'Tarde'
        else:
            return 'Noche'
        
    def _cumple_restricciones_tiempo(self, profesor_id, dia, hora_inicio, hora_fin, horario_generado, params):
        """Verifica si se cumplen las restricciones de tiempo consecutivo"""
        # Convertir horarios a minutos para facilitar comparaciones
        inicio_actual = hora_inicio.hour * 60 + hora_inicio.minute
        fin_actual = hora_fin.hour * 60 + hora_fin.minute
        
        # Obtener clases del profesor en el mismo día
        clases_dia = [
            clase for clase in horario_generado
            if clase['profesor_id'] == profesor_id and clase['dia_semana'] == dia
        ]
        
        # Verificar máximo de horas por día
        horas_dia = sum(
            self._calcular_duracion_minutos(clase['hora_inicio'], clase['hora_fin'])
            for clase in clases_dia
        ) / 60
        
        if horas_dia + 1.5 > params['max_horas_dia']:  # 1.5 horas = 90 minutos
            return False
        
        # Verificar tiempo mínimo de descanso
        for clase in clases_dia:
            fin_clase = self.parse_time(clase['hora_fin'])
            inicio_clase = self.parse_time(clase['hora_inicio'])
            
            fin_clase_min = fin_clase.hour * 60 + fin_clase.minute
            inicio_clase_min = inicio_clase.hour * 60 + inicio_clase.minute
            
            # Verificar si hay suficiente descanso entre clases
            if (abs(inicio_actual - fin_clase_min) < params['min_descanso'] or
                abs(fin_actual - inicio_clase_min) < params['min_descanso']):
                return False
        
        return True
    
    def _es_hora_valida_para_jornada(self, hora, preferencias, bloques_asignados):
        map_jornadas = {
            'Mañana': (6, 13),
            'Tarde': (13, 18), 
            'Noche': (18, 23)
        }
        
        bloques_ideales = len(preferencias) > 0 and sum(bloques_asignados.values()) // len(preferencias)
        
        for jornada, (inicio, fin) in map_jornadas.items():
            if jornada in preferencias and inicio <= hora.hour < fin:
                return bloques_asignados[jornada] < bloques_ideales or sum(bloques_asignados.values()) == 0, jornada
                
        return False, None

    def _calcular_duracion_minutos(self, hora_inicio, hora_fin):
        """Calcula la duración en minutos entre dos horas"""
        inicio = self.parse_time(hora_inicio)
        fin = self.parse_time(hora_fin)
        
        inicio_min = inicio.hour * 60 + inicio.minute
        fin_min = fin.hour * 60 + fin.minute
        
        return fin_min - inicio_min

    def _analizar_disponibilidad_jornadas(self, materia, df_profesores, df_horarios, df_profesor_materia):
        jornadas_disponibles = []
        
        for jornada in ['mañana', 'tarde', 'noche']:
            df_horarios_jornada = self._filtrar_horarios_por_jornada(df_horarios, jornada)
            profesores_disponibles = df_profesor_materia[df_profesor_materia['materia_id'] == materia['id']]['profesor_id'].unique()
            
            disponibilidad = df_horarios_jornada[df_horarios_jornada['profesor_id'].isin(profesores_disponibles)].groupby('profesor_id').size()
            
            if len(disponibilidad[disponibilidad >= materia['bloques'] * 2]) > 0:
                jornadas_disponibles.append(jornada)
        
        return jornadas_disponibles
    
    def _filtrar_horarios_por_jornada(self, df_horarios, jornada):
        """Filtra los horarios según la jornada especificada y distribuye aleatoriamente"""
        rangos = {
            'mañana': (self.parse_time('06:00'), self.parse_time('12:59')),
            'tarde': (self.parse_time('13:00'), self.parse_time('17:59')),
            'noche': (self.parse_time('18:00'), self.parse_time('22:00'))
        }
        
        inicio, fin = rangos[jornada]
        horarios_jornada = df_horarios[
            df_horarios['hora_inicio'].apply(self.parse_time).between(inicio, fin)
        ]
        
        if not horarios_jornada.empty:
            # Aleatorizar las horas dentro de la jornada
            horas_permitidas = self.HORAS_PERMITIDAS[jornada]
            horarios_jornada['hora_inicio'] = horarios_jornada.apply(
                lambda x: random.choice(horas_permitidas),
                axis=1
            )
            
            # Ajustar hora_fin basado en la nueva hora_inicio
            horarios_jornada['hora_fin'] = horarios_jornada['hora_inicio'].apply(
                lambda x: (datetime.strptime(x, '%H:%M') + timedelta(minutes=45)).strftime('%H:%M')
            )
        
        return horarios_jornada

    def _obtener_profesores_elegibles_jornada(self, materia, df_profesores, df_profesor_materia, 
                                            df_horarios_jornada, horario_generado, params):
        """Obtiene lista de profesores elegibles con disponibilidad en la jornada específica"""
        profesores_elegibles = []
        for _, profesor in df_profesores.iterrows():
            if (profesor['estado'] == 'Activo' and
                self.get_carga_actual(profesor['id'], horario_generado) < params['max_carga_profesor']):
                
                # Verificar si el profesor tiene la materia asignada
                prof_mat = df_profesor_materia[
                    (df_profesor_materia['profesor_id'] == profesor['id']) &
                    (df_profesor_materia['materia_id'] == materia['id'])
                ]
                
                if not prof_mat.empty:
                    # Verificar disponibilidad en la jornada
                    horarios_prof = df_horarios_jornada[
                        df_horarios_jornada['profesor_id'] == profesor['id']
                    ]
                    
                    if len(horarios_prof) >= materia['bloques'] * 2:
                        profesores_elegibles.append(profesor.to_dict())
        
        return profesores_elegibles
            
    def _preprocesar_horarios(self, df_horarios, params):
        """Preprocesa y filtra horarios según parámetros"""
        hora_inicio = self.parse_time(params['horario_inicio'])
        hora_fin = self.parse_time(params['horario_fin'])
        
        df_filtrado = df_horarios[
            (df_horarios['dia'].isin(params['dias_habiles'])) &
            (df_horarios['hora_inicio'].apply(self.parse_time) >= hora_inicio) &
            (df_horarios['hora_fin'].apply(self.parse_time) <= hora_fin)
        ].copy()
        
        return df_filtrado
            

    def _obtener_profesores_elegibles(self, materia, df_profesores, df_profesor_materia, horario_generado, params):
        """Obtiene lista de profesores elegibles aplicando los criterios disponibles"""
        profesores_elegibles = []
        for _, profesor in df_profesores.iterrows():
            # Obtener información de profesor_materia
            prof_mat = df_profesor_materia[
                (df_profesor_materia['profesor_id'] == profesor['id']) &
                (df_profesor_materia['materia_id'] == materia['id'])
            ]
            
            if (profesor['estado'] == 'Activo' and
                self.get_carga_actual(profesor['id'], horario_generado) < params['max_carga_profesor'] and
                not prof_mat.empty and
                prof_mat['experiencia'].iloc[0] >= params.get('min_experiencia', 0) and
                prof_mat['calificacion_alumno'].iloc[0] >= params.get('min_calificacion', 0)):
                
                profesores_elegibles.append(profesor.to_dict())
        
        return profesores_elegibles

    def _crear_indice_horarios_ocupados(self, horario_generado):
        """Crea un índice de horarios ocupados para búsqueda rápida"""
        horarios_ocupados = {}
        for clase in horario_generado:
            key = (clase['dia_semana'], clase['hora_inicio'], clase['hora_fin'])
            horarios_ocupados[key] = True
        return horarios_ocupados

    def _encontrar_mejor_asignacion_rapida(self, materia, profesores_elegibles, salones_adecuados, df_horarios_filtrado, dias_asignados, horarios_ocupados, params):
        # Crear índices hash para búsquedas O(1)
        horarios_dict = {profesor['id']: df_horarios_filtrado[df_horarios_filtrado['profesor_id'] == profesor['id']] 
                        for profesor in profesores_elegibles}
        
        dias_disponibles = set(params['dias_habiles']) - dias_asignados if dias_asignados else set(params['dias_habiles'])
        
        mejor_score = -1
        mejor_asignacion = None
        
        # Búsqueda por batches 
        for profesor in profesores_elegibles:
            if not horarios_dict[profesor['id']].empty:
                for _, horario in horarios_dict[profesor['id']].iterrows():
                    # Vectorización de cálculos
                    scores = np.array([self._calcular_score_rapido(profesor, salon, horario, materia) 
                                    for _, salon in salones_adecuados.iterrows()])
                    max_score_idx = np.argmax(scores)
                    
                    if scores[max_score_idx] > mejor_score:
                        mejor_score = scores[max_score_idx]
                        mejor_asignacion = (profesor, salones_adecuados.iloc[max_score_idx])
                        
                        if mejor_score > 0.8:
                            return mejor_asignacion
                            
        return mejor_asignacion

    @st.cache_data
    def _calcular_score_rapido(self, profesor, salon, horario, materia):
        score = 0.5
        ratio_capacidad = salon['capacidad_alumnos'] / materia['alumnos']
        
        # Usar lookup tables para decisiones comunes
        SCORE_MODS = {
            (1.0, 1.5): 0.2,
            (1.5, float('inf')): 0.1
        }
        
        for (min_ratio, max_ratio), mod in SCORE_MODS.items():
            if min_ratio <= ratio_capacidad < max_ratio:
                score += mod
                break
                
        return min(1.0, score)
    
    def _get_grupo_prefix(self, hora_inicio):
        """
        Determina el prefijo del grupo según la hora del día
        
        Args:
            hora_inicio: string u objeto time con la hora de inicio
        
        Returns:
            str: Prefijo para el código del grupo
        """
        if isinstance(hora_inicio, str):
            hora = datetime.strptime(hora_inicio, '%H:%M').time()
        else:
            hora = hora_inicio
        
        # Mañana: 6:00 - 12:59
        if hora.hour >= 6 and hora.hour < 13:
            return 'B'
        # Tarde: 13:00 - 17:59
        elif hora.hour >= 13 and hora.hour < 18:
            return 'C'
        # Noche: 18:00 - 22:00
        else:
            return 'E'

    def _crear_bloques_clase(self, materia, profesor, salon, horario, num_bloques):
        """
        Crea los bloques de clase asegurando que cada bloque sea de 90 minutos (2 clases de 45)
        """
        bloques = []
        dia = horario['dia']
        
        # Determinar jornada
        jornada = self._get_jornada_from_hora(horario['hora_inicio'])
        if not jornada:
            return bloques
            
        # Obtener o crear grupo para este profesor y materia
        grupo_codigo = self._asignar_grupo_profesor(
            materia['id'],
            profesor['id'],
            jornada
        )
        
        hora = self.parse_time(horario['hora_inicio'])
        
        # Determinar límite de la jornada
        _, limite_fin = self.JORNADAS[jornada]
        
        # Crear el número especificado de clases de 45 minutos
        for i in range(num_bloques):
            hora_inicio_bloque = (datetime.combine(datetime.min, hora) + 
                                timedelta(minutes=45*i)).time()
            hora_fin_bloque = (datetime.combine(datetime.min, hora_inicio_bloque) + 
                            timedelta(minutes=45)).time()
            
            if hora_fin_bloque <= limite_fin:
                bloque = {
                    'grupo': grupo_codigo,
                    'dia_semana': dia,
                    'hora_inicio': hora_inicio_bloque.strftime('%H:%M'),
                    'hora_fin': hora_fin_bloque.strftime('%H:%M'),
                    'alumnos': int(materia['alumnos']),
                    'materia_id': int(materia['id']),
                    'salon_id': int(salon['id']),
                    'profesor_id': int(profesor['id'])
                }
                bloques.append(bloque)
        
        return bloques if len(bloques) == num_bloques else []

    def _actualizar_profesores_elegibles(self, profesores_elegibles, profesor_id, horario_generado, params):
        """Actualiza la lista de profesores elegibles"""
        for i, profesor in enumerate(profesores_elegibles):
            if (profesor['id'] == profesor_id and 
                self.get_carga_actual(profesor_id, horario_generado) >= params['max_carga_profesor']):
                profesores_elegibles.pop(i)
                break

    def encontrar_mejor_asignacion(self, materia, profesores, salones, horarios, params):
            """Encuentra la mejor asignación considerando múltiples factores"""
            mejor_score = -1
            mejor_asignacion = None
            
            for profesor in profesores:
                if self.get_carga_actual(profesor['id'], []) >= params['max_carga_profesor']:
                    continue
                    
                horarios_prof = horarios[horarios['profesor_id'] == profesor['id']]
                if horarios_prof.empty:
                    continue
                    
                for _, horario in horarios_prof.iterrows():
                    for _, salon in salones.iterrows():
                        if salon['capacidad_alumnos'] < materia['alumnos']:
                            continue
                            
                        score = self._calcular_score_rapido(profesor, salon, horario, materia)
                        if score > mejor_score:
                            mejor_score = score
                            mejor_asignacion = (profesor, salon, horario, score)
                            if score > 0.8:  # Early exit si encontramos buena asignación
                                return mejor_asignacion
                                
            return mejor_asignacion

    def parse_time(self, time_str):
        """Convierte una cadena de tiempo en un objeto time"""
        #esta cosa me mato como 4 semanas sebastian, pero funcionó
        if isinstance(time_str, (datetime, time)):
            return time_str.time() if isinstance(time_str, datetime) else time_str
        try:
            return datetime.strptime(time_str, '%H:%M:%S').time()
        except ValueError:
            return datetime.strptime(time_str, '%H:%M').time()


    def get_carga_actual(self, profesor_id, horario_generado):
        return len([
            clase for clase in horario_generado 
            if clase['profesor_id'] == profesor_id
        ])

    def hay_conflicto_horario(self, profesor_id, horario_nuevo, horario_generado):
        for clase in horario_generado:
            if (clase['profesor_id'] == profesor_id and
                clase['dia_semana'] == horario_nuevo['dia'] and
                self.hay_solapamiento(
                    self.parse_time(clase['hora_inicio']),
                    self.parse_time(clase['hora_fin']),
                    horario_nuevo['hora_inicio'],
                    horario_nuevo['hora_fin']
                )):
                return True
        return False

    def salon_ocupado(self, salon_id, horario_nuevo, horario_generado):
        for clase in horario_generado:
            if (clase['salon_id'] == salon_id and
                clase['dia_semana'] == horario_nuevo['dia'] and
                self.hay_solapamiento(
                    clase['hora_inicio'], clase['hora_fin'],
                    horario_nuevo['hora_inicio'], horario_nuevo['hora_fin']
                )):
                return True
        return False

def main():
    st.title('🎓Proyecto de grado: Sistema Avanzado de Generacion de clases de las UTS')
    
    # Menú de navegación
    selected = option_menu(
        menu_title=None,
        options=["Dashboard", "Configuración", "Entrenamiento", "Generación", "Análisis"],
        icons=["house", "gear", "book", "calendar", "graph-up"],
        menu_icon="cast",
        default_index=0,
        orientation="horizontal",
    )
    
    optimizer = ScheduleOptimizer()
    
    if selected == "Dashboard":
        show_dashboard(optimizer)
    elif selected == "Configuración":
        show_configuration(optimizer)
    elif selected == "Entrenamiento":
        show_training(optimizer)
    elif selected == "Generación":
        show_generation(optimizer)
    elif selected == "Análisis":
        show_analysis(optimizer)

def show_dashboard(optimizer):
    st.header("📊 Dashboard General")
    
    # Métricas principales en una fila
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            label="Modelo Activo",
            value="KNN Y Random Forest" if optimizer.is_fitted else "No entrenado",
            delta="Activo y Funcionando" if optimizer.is_fitted else None,
            delta_color="normal" if optimizer.is_fitted else "off"
        )
    
    with col2:
        precision = f"{optimizer.best_model_score:.2%}" if hasattr(optimizer, 'best_model_score') else "N/A"
        st.metric(
            label="Precisión del Modelo",
            value=precision,
            help="Precisión del último entrenamiento del modelo"
        )
    
    with col3:
        num_horarios = len(optimizer.horarios_generados)
        ultimo_horario = optimizer.horarios_generados[-1] if optimizer.horarios_generados else None
        delta = "Último: " + ultimo_horario['status'] if ultimo_horario else None
        
        st.metric(
            label="Horarios Generados",
            value=num_horarios,
            delta=delta,
            help="Número total de horarios generados exitosamente"
        )
    
    with col4:
        if optimizer.horarios_generados:
            success_rate = len([h for h in optimizer.horarios_generados if h['status'] == 'OPTIMAL']) / len(optimizer.horarios_generados)
            st.metric(
                label="Tasa de Éxito",
                value=f"{success_rate:.2%}",
                help="Porcentaje de horarios generados sin conflictos"
            )
        else:
            st.metric(label="Tasa de Éxito", value="N/A")
    
    # Gráfico de generación de horarios a lo largo del tiempo
    if optimizer.horarios_generados:
        st.subheader("📈 Historial de Generación de Horarios")
        
        df_historico = pd.DataFrame(optimizer.horarios_generados)
        df_historico['fecha'] = pd.to_datetime(df_historico['fecha'])
        
        fig_historico = px.line(
            df_historico,
            x='fecha',
            y='num_clases',
            title='Clases por Horario Generado',
            labels={'fecha': 'Fecha de Generación', 'num_clases': 'Número de Clases'}
        )
        fig_historico.update_traces(mode='lines+markers')
        st.plotly_chart(fig_historico, use_container_width=True)
        
        # Estadísticas adicionales
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("📊 Estadísticas de Generación")
            stats_df = pd.DataFrame({
                'Métrica': ['Total Horarios', 'Horarios Óptimos', 'Horarios con Advertencias'],
                'Valor': [
                    len(df_historico),
                    len(df_historico[df_historico['status'] == 'OPTIMAL']),
                    len(df_historico[df_historico['warnings'] > 0])
                ]
            })
            st.dataframe(stats_df, hide_index=True)
        
        with col2:
            st.subheader("🎯 Últimos Horarios Generados")
            recent_df = df_historico.tail(5)[['fecha', 'status', 'num_clases', 'warnings']]
            st.dataframe(recent_df, hide_index=True)
    
    else:
        st.info("👋 Aún no se han generado horarios. Dirígete a la sección de Generación para crear tu primer horario.")


def show_configuration(optimizer):
    st.header("⚙️ Configuración del Sistema")
    
    # Intentar cargar configuración existente
    saved_config = optimizer.load_configuration()
    
    # Crear pestañas para diferentes secciones de configuración
    tabs = st.tabs([
        "Parámetros Básicos", 
        "Configuración Avanzada", 
        "Agente Adaptativo",
        "Restricciones del Sistema"
    ])
    
    # Parámetros Básicos
    with tabs[0]:
        st.subheader("🎯 Parámetros Básicos")
        
        col1, col2 = st.columns(2)
        with col1:
            basic_params = {
                'slot_duration': st.slider(
                    "Duración del slot (minutos)", 
                    min_value=30,
                    max_value=120,
                    value=saved_config.get('basic', {}).get('slot_duration', optimizer.slot_duration) if saved_config else optimizer.slot_duration,
                    step=15,
                    help="Duración de cada bloque de tiempo para las clases"
                ),
                'min_alumnos': st.number_input(
                    "Mínimo de alumnos por clase",
                    min_value=1,
                    value=saved_config.get('basic', {}).get('min_alumnos', 10) if saved_config else 10,
                    help="Número mínimo de alumnos requeridos para abrir una clase"
                ),
                'max_carga_profesor': st.number_input(
                    "Carga máxima profesor",
                    min_value=1,
                    max_value=40,
                    value=saved_config.get('basic', {}).get('max_carga_profesor', 20) if saved_config else 20,
                    help="Número máximo de horas que puede dar un profesor"
                )
            }
        
        with col2:
            dias_default = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
            basic_params.update({
                'dias_habiles': st.multiselect(
                    "Días hábiles",
                    options=["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"],
                    default=saved_config.get('basic', {}).get('dias_habiles', dias_default) if saved_config else dias_default,
                    help="Días disponibles para programar clases"
                ),
                'horario_inicio': st.time_input(
                    "Hora de inicio de jornada",
                    value=datetime.strptime(
                        saved_config.get('basic', {}).get('horario_inicio', "06:00") if saved_config else "06:00",
                        "%H:%M"
                    ).time(),
                    help="Hora de inicio de la jornada académica"
                ),
                'horario_fin': st.time_input(
                    "Hora de fin de jornada",
                    value=datetime.strptime(
                        saved_config.get('basic', {}).get('horario_fin', "22:00") if saved_config else "22:00",
                        "%H:%M"
                    ).time(),
                    help="Hora de finalización de la jornada académica"
                )
            })
            
            # Validación de horarios
            if basic_params['horario_inicio'] >= basic_params['horario_fin']:
                st.error("❌ La hora de inicio debe ser anterior a la hora de fin")
    
    # Configuración Avanzada
    with tabs[1]:
        st.subheader("🔧 Configuración Avanzada")
        
        col1, col2 = st.columns(2)
        with col1:
            advanced_params = {
                'optimization_level': st.select_slider(
                    "Nivel de optimización",
                    options=["Bajo", "Medio", "Alto"],
                    value=saved_config.get('advanced', {}).get('optimization_level', "Medio") if saved_config else "Medio",
                    help="Define qué tan exhaustiva será la búsqueda de soluciones óptimas"
                ),
                'max_iterations': st.number_input(
                    "Máximo de iteraciones",
                    min_value=100,
                    max_value=10000,
                    value=saved_config.get('advanced', {}).get('max_iterations', 1000) if saved_config else 1000,
                    step=100,
                    help="Número máximo de iteraciones para buscar solución"
                )
            }
        
        with col2:
            advanced_params.update({
                'allow_overlap': st.checkbox(
                    "Permitir solapamientos controlados",
                    value=saved_config.get('advanced', {}).get('allow_overlap', False) if saved_config else False,
                    help="Permite solapamientos bajo ciertas condiciones"
                ),
                'priority_rules': st.multiselect(
                    "Reglas de prioridad",
                    options=[
                        "Materias con más alumnos",
                        "Profesores más experimentados",
                        "Salones más grandes primero",
                        "Horarios más temprano"
                    ],
                    default=saved_config.get('advanced', {}).get('priority_rules', ["Materias con más alumnos"]) if saved_config else ["Materias con más alumnos"],
                    help="Reglas para priorizar la asignación de recursos"
                )
            })

    # Agente Adaptativo
    with tabs[2]:
        st.subheader("🤖 Configuración del Agente Adaptativo")
        
        col1, col2 = st.columns(2)
        with col1:
            adaptive_params = {
                'learning_rate': st.slider(
                    "Tasa de aprendizaje",
                    min_value=0.01,
                    max_value=1.0,
                    value=saved_config.get('adaptive', {}).get('learning_rate', optimizer.adaptive_agent.learning_rate) if saved_config else optimizer.adaptive_agent.learning_rate,
                    format="%.2f",
                    help="Velocidad de adaptación del agente"
                ),
                'adaptation_threshold': st.slider(
                    "Umbral de adaptación",
                    min_value=0.0,
                    max_value=1.0,
                    value=saved_config.get('adaptive', {}).get('adaptation_threshold', optimizer.adaptive_agent.adaptation_threshold) if saved_config else optimizer.adaptive_agent.adaptation_threshold,
                    format="%.2f",
                    help="Umbral para activar adaptaciones"
                )
            }
        
        with col2:
            adaptive_params.update({
                'enable_pattern_detection': st.checkbox(
                    "Habilitar detección de patrones",
                    value=saved_config.get('adaptive', {}).get('enable_pattern_detection', True) if saved_config else True,
                    help="Permite al sistema aprender de patrones exitosos"
                ),
                'pattern_memory_size': st.number_input(
                    "Tamaño de memoria de patrones",
                    min_value=10,
                    max_value=1000,
                    value=saved_config.get('adaptive', {}).get('pattern_memory_size', 100) if saved_config else 100,
                    help="Número máximo de patrones a recordar"
                )
            })
        
        # Métricas del agente adaptativo si están disponibles
        if hasattr(optimizer.adaptive_agent, 'get_performance_metrics'):
            metrics = optimizer.adaptive_agent.get_performance_metrics()
            if metrics:
                st.subheader("📊 Métricas del Agente Adaptativo")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Tasa de aprendizaje actual", f"{metrics['current_learning_rate']:.3f}")
                with col2:
                    st.metric("Rendimiento promedio", f"{metrics['average_performance']:.2%}")
                with col3:
                    st.metric("Tendencia de mejora", f"{metrics['improvement_trend']:.2%}")

    # Restricciones del Sistema
    with tabs[3]:
        st.subheader("🚫 Restricciones del Sistema")
        
        col1, col2 = st.columns(2)
        with col1:
            restriction_params = {
                'max_clases_consecutivas': st.number_input(
                    "Máximo de clases consecutivas",
                    min_value=1,
                    max_value=6,
                    value=saved_config.get('restrictions', {}).get('max_clases_consecutivas', 3) if saved_config else 3,
                    help="Número máximo de clases consecutivas permitidas"
                ),
                'min_descanso': st.number_input(
                    "Tiempo mínimo de descanso (minutos)",
                    min_value=0,
                    max_value=60,
                    value=saved_config.get('restrictions', {}).get('min_descanso', 15) if saved_config else 15,
                    help="Tiempo mínimo de descanso entre clases"
                )
            }
        
        with col2:
            restriction_params.update({
                'max_ventanas': st.number_input(
                    "Máximo de ventanas por día",
                    min_value=0,
                    max_value=5,
                    value=saved_config.get('restrictions', {}).get('max_ventanas', 2) if saved_config else 2,
                    help="Número máximo de períodos libres entre clases"
                ),
                'distancia_maxima': st.number_input(
                    "Distancia máxima entre salones (metros)",
                    min_value=0,
                    max_value=1000,
                    value=saved_config.get('restrictions', {}).get('distancia_maxima', 100) if saved_config else 100,
                    help="Distancia máxima permitida entre salones consecutivos"
                )
            })

    # Contenedor para botones de acción
    button_col1, button_col2 = st.columns([2, 1])
    
    with button_col1:
        # Botón para guardar la configuración
        if st.button("💾 Guardar Configuración", type="primary", use_container_width=True):
            # Combinar todos los parámetros
            config = {
                "basic": basic_params,
                "advanced": advanced_params,
                "adaptive": adaptive_params,
                "restrictions": restriction_params
            }
            
            # Intentar guardar la configuración
            success, message = optimizer.save_configuration(config)
            
            if success:
                st.success(f"✅ {message}")
                
                # Mostrar resumen de cambios
                with st.expander("📋 Resumen de la configuración guardada"):
                    for section, params in config.items():
                        st.subheader(section.title())
                        # Crear un DataFrame para mejor visualización
                        df = pd.DataFrame(list(params.items()), columns=['Parámetro', 'Valor'])
                        st.dataframe(df, hide_index=True)
                        
                # Mostrar información sobre la ubicación del archivo
                st.info("📁 Configuración guardada en: config/system_config.json")
            else:
                st.error(f"❌ {message}")
                with st.expander("🔍 Detalles técnicos del error"):
                    st.code(message)
    
    with button_col2:
        # Botón para resetear la configuración
        if st.button("🔄 Resetear valores", type="secondary", use_container_width=True):
            if os.path.exists('config/system_config.json'):
                try:
                    os.remove('config/system_config.json')
                    st.success("✅ Configuración reseteada a valores predeterminados")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error al resetear la configuración: {str(e)}")

    # Mostrar advertencias si hay configuraciones potencialmente problemáticas
    if basic_params['slot_duration'] < 45:
        st.warning("⚠️ Una duración de slot menor a 45 minutos podría no ser óptima para clases regulares")
    
    if advanced_params['allow_overlap']:
        st.warning("⚠️ Permitir solapamientos puede generar conflictos en los horarios")
    
    if restriction_params['max_clases_consecutivas'] > 4:
        st.warning("⚠️ Un número alto de clases consecutivas podría afectar el rendimiento de profesores y estudiantes")

    # Mostrar recomendaciones basadas en la configuración actual
    with st.expander("💡 Recomendaciones de optimización"):
        st.markdown("""
        - **Duración de slots**: Para mejor aprovechamiento, considera slots de 45-90 minutos
        - **Carga docente**: Distribuye la carga de manera equilibrada entre profesores
        - **Ventanas**: Minimiza los espacios libres entre clases
        - **Patrones**: Habilita la detección de patrones para mejorar la optimización
        """)

def show_training(optimizer):
    st.header("🎯 Entrenamiento del Modelo")
    
    # Crear dos columnas para organizar los parámetros
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("Configuración del Modelo")
        
        # Parámetros básicos del modelo
        model_type = st.selectbox(
            "Tipo de Modelo",
            ["Random Forest", "KNN"],
            help="Seleccione el algoritmo de aprendizaje automático a utilizar"
        )
        
        # Parámetros específicos según el tipo de modelo
        model_params = {}
        if model_type == "Random Forest":
            col_rf1, col_rf2 = st.columns(2)
            with col_rf1:
                model_params['n_estimators'] = st.slider(
                    "Número de árboles",
                    min_value=50,
                    max_value=500,
                    value=100,
                    step=50,
                    help="Cantidad de árboles en el bosque aleatorio"
                )
                model_params['max_depth'] = st.slider(
                    "Profundidad máxima",
                    min_value=3,
                    max_value=50,
                    value=10,
                    help="Profundidad máxima de cada árbol"
                )
            with col_rf2:
                model_params['min_samples_split'] = st.slider(
                    "Muestras mínimas para división",
                    min_value=2,
                    max_value=10,
                    value=2,
                    help="Número mínimo de muestras requeridas para dividir un nodo"
                )
                model_params['min_samples_leaf'] = st.slider(
                    "Muestras mínimas en hojas",
                    min_value=1,
                    max_value=10,
                    value=1,
                    help="Número mínimo de muestras requeridas en un nodo hoja"
                )
        else:  # KNN
            col_knn1, col_knn2 = st.columns(2)
            with col_knn1:
                model_params['n_neighbors'] = st.slider(
                    "Número de vecinos (K)",
                    min_value=1,
                    max_value=20,
                    value=5,
                    help="Número de vecinos a considerar"
                )
            with col_knn2:
                model_params['weights'] = st.selectbox(
                    "Ponderación",
                    ['uniform', 'distance'],
                    help="Método de ponderación de los vecinos"
                )
                model_params['metric'] = st.selectbox(
                    "Métrica de distancia",
                    ['euclidean', 'manhattan'],
                    help="Métrica para calcular la distancia entre puntos"
                )

        # Parámetros generales de entrenamiento
        st.subheader("Parámetros de Entrenamiento")
        col_gen1, col_gen2 = st.columns(2)
        with col_gen1:
            test_size = st.slider(
                "Tamaño del conjunto de prueba",
                min_value=0.1,
                max_value=0.4,
                value=0.2,
                help="Proporción de datos para prueba"
            )
            model_params['test_size'] = test_size
            
        with col_gen2:
            cv_folds = st.slider(
                "Folds de validación cruzada",
                min_value=2,
                max_value=10,
                value=5,
                help="Número de particiones para validación cruzada"
            )
            model_params['cv_folds'] = cv_folds

    with col2:
        st.subheader("Estado del Entrenamiento")
        
        # Métricas del modelo actual
        if optimizer.is_fitted:
            st.success("Modelo entrenado ✓")
            if hasattr(optimizer, 'best_model_score'):
                st.metric("Precisión actual", f"{optimizer.best_model_score:.2%}")
            else:
                st.metric("Precisión actual", "No disponible")
                
            if hasattr(optimizer, 'last_training_date'):
                st.metric("Última actualización", 
                         optimizer.last_training_date if optimizer.last_training_date else "No disponible")
            else:
                st.metric("Última actualización", "No disponible")
        else:
            st.warning("Modelo no entrenado")
            
    # Botón de entrenamiento
    if st.button("🚀 Iniciar Entrenamiento", use_container_width=True):
        with st.spinner("Entrenando modelo..."):
            try:
                # Preparar datos
                data = {
                    'profesores': optimizer.get_data('profesores'),
                    'materias': optimizer.get_data('materias'),
                    'salones': optimizer.get_data('salones'),
                    'horarios_disponibles': optimizer.get_data('horarios_disponibles'),
                    'profesor_materia': optimizer.get_data('profesor_materia')
                }
                
                if all(data.values()):
                    dfs = {k: pd.DataFrame(v) for k, v in data.items()}
                    
                    # Preparar características
                    X, y, conflicts = optimizer.prepare_features(
                        dfs['profesores'],
                        dfs['materias'],
                        dfs['salones'],
                        dfs['horarios_disponibles'],
                        dfs['profesor_materia']
                    )
                    
                    # Entrenar modelo
                    results = optimizer.train_model(X, y, model_params)
                    
                    # Mostrar resultados
                    st.success("✅ Modelo entrenado exitosamente")
                    
                    # Crear tabs para mostrar diferentes métricas
                    metric_tabs = st.tabs(["Métricas Generales", "Matriz de Confusión", "Importancia de Características"])
                    
                    with metric_tabs[0]:
                        st.subheader("Métricas de Clasificación")
                        report_df = pd.DataFrame(results['classification_report']).transpose()
                        st.dataframe(report_df.style.highlight_max(axis=0))
                        
                    with metric_tabs[1]:
                        st.subheader("Matriz de Confusión")
                        conf_matrix = results['confusion_matrix']
                        fig = go.Figure(data=go.Heatmap(
                            z=conf_matrix,
                            x=['Negativo', 'Positivo'],
                            y=['Negativo', 'Positivo'],
                            text=conf_matrix,
                            texttemplate="%{text}",
                            textfont={"size": 16},
                            hoverongaps=False,
                            colorscale='Blues'
                        ))
                        
                        fig.update_layout(
                            title='Matriz de Confusión',
                            xaxis_title='Predicho',
                            yaxis_title='Real',
                            width=500,
                            height=500
                        )
                        
                        st.plotly_chart(fig)
                        
                    with metric_tabs[2]:
                        if 'feature_importance' in results:
                            st.subheader("Importancia de Características")
                            importance_df = pd.DataFrame({
                                'Característica': list(results['feature_importance'].keys()),
                                'Importancia': list(results['feature_importance'].values())
                            }).sort_values('Importancia', ascending=False)
                            
                            fig = px.bar(
                                importance_df,
                                x='Característica',
                                y='Importancia',
                                title='Importancia de Características'
                            )
                            st.plotly_chart(fig)
                            
                else:
                    st.error("No se pudieron cargar todos los datos necesarios")
            except Exception as e:
                st.error(f"Error durante el entrenamiento: {str(e)}")

def show_generation(optimizer):
    st.header("📅 Generacion de Horarios")
    
    if not optimizer.is_fitted:
        st.warning("⚠️ El modelo no está entrenado. Por favor, entrene el modelo primero.")
        if st.button("Ir a Entrenamiento"):
            st.session_state.page = "training"
        return

    if 'horario_generado' not in st.session_state:
        st.session_state.horario_generado = None
        st.session_state.resultado = None
        st.session_state.dfs = None

    col_main, col_sidebar = st.columns([3, 1])

    with col_main:
        st.subheader("📊 Configuracion de Generacion de clases")
        
        tab_basic, tab_advanced, tab_restrictions = st.tabs([
            "Parámetros Básicos", 
            "Configuración Avanzada", 
            "Restricciones"
        ])

        with tab_basic:
            col1, col2 = st.columns(2)
            with col1:
                optimization_params = {
                    'min_alumnos': st.number_input(
                        "Mínimo de alumnos por clase",
                        min_value=1,
                        value=10,
                        help="Número mínimo de alumnos para abrir una clase"
                    ),
                    'max_carga_profesor': st.number_input(
                        "Máxima carga por profesor",
                        min_value=1,
                        value=20,
                        help="Número máximo de horas que puede dar un profesor"
                    )
                }
            
            with col2:
                optimization_params.update({
                    'min_experiencia': st.number_input(
                        "Experiencia mínima requerida",
                        min_value=0,
                        value=1,
                        help="Años mínimos de experiencia requeridos"
                    ),
                    'min_calificacion': st.number_input(
                        "Calificación mínima del profesor",
                        min_value=1.0,
                        max_value=5.0,
                        value=3.0,
                        step=0.1,
                        help="Calificación mínima aceptable del profesor"
                    )
                })

        with tab_advanced:
            col1, col2 = st.columns(2)
            with col1:
                optimization_params.update({
                    'optimization_level': st.select_slider(
                        'Nivel de optimización',
                        options=['Bajo', 'Medio', 'Alto'],
                        value='Medio',
                        help="Define qué tan exhaustiva será la búsqueda de soluciones óptimas"
                    ),
                    'conflict_tolerance': st.slider(
                        'Tolerancia a conflictos',
                        min_value=0.0,
                        max_value=1.0,
                        value=0.1,
                        help="Nivel de tolerancia para conflictos en la generación de horarios"
                    )
                })
            
            with col2:
                optimization_params.update({
                    'enable_pattern_detection': st.checkbox(
                        'Habilitar detección de patrones',
                        value=True,
                        help="Permite al sistema aprender de patrones exitosos anteriores"
                    ),
                    'auto_correction': st.checkbox(
                        'Habilitar auto-corrección',
                        value=True,
                        help="Permite al sistema corregir automáticamente conflictos menores"
                    )
                })

        with tab_restrictions:
            col1, col2 = st.columns(2)
            with col1:
                optimization_params.update({
                    'max_dias_consecutivos': st.number_input(
                        "Máximo de días consecutivos",
                        min_value=1,
                        max_value=6,
                        value=5,
                        help="Máximo de días consecutivos que un profesor puede dar clases"
                    ),
                    'max_horas_dia': st.number_input(
                        "Máximo de horas por día",
                        min_value=1,
                        max_value=12,
                        value=8,
                        help="Máximo de horas que un profesor puede dar en un día"
                    )
                })
            
            with col2:
                optimization_params.update({
                    'min_descanso': st.number_input(
                        "Mínimo de descanso (minutos)",
                        min_value=0,
                        max_value=120,
                        value=30,
                        step=15,
                        help="Tiempo mínimo de descanso entre clases"
                    ),
                    'preferencia_horario': st.multiselect(
                        "Preferencia de horario",
                        options=["Mañana", "Tarde", "Noche"],
                        default=["Mañana", "Tarde"],
                        help="Preferencias de horario para la asignación"
                    )
                })

    with col_sidebar:
        st.subheader("🎯 Estado de Generación")
        if optimizer.is_fitted:
            st.success("Modelo listo para generar")
            st.metric("Precisión del modelo", f"{optimizer.best_model_score:.2%}")
            if hasattr(optimizer, 'last_generation_time'):
                st.metric("Última generación", optimizer.last_generation_time)

    generar_horario = st.button("🎲 Generar Horario", type="primary", use_container_width=True)
    
    if generar_horario:
        with st.spinner("Generando horario optimizado..."):
            try:
                data = {
                    'profesores': optimizer.get_data('profesores'),
                    'materias': optimizer.get_data('materias'),
                    'salones': optimizer.get_data('salones'),
                    'horarios_disponibles': optimizer.get_data('horarios_disponibles'),
                    'profesor_materia': optimizer.get_data('profesor_materia')
                }
                
                if all(data.values()):
                    dfs = {k: pd.DataFrame(v) for k, v in data.items()}
                    resultado = optimizer.generate_schedule(
                        dfs['profesores'],
                        dfs['materias'],
                        dfs['salones'],
                        dfs['horarios_disponibles'],
                        dfs['profesor_materia'],
                        optimization_params
                    )
                    st.session_state.resultado = resultado
                    st.session_state.dfs = dfs
                    st.rerun()

            except Exception as e:
                st.error(f"❌ Error durante la generación del horario: {str(e)}")
                return

    # Mostrar resultados si existen
    if hasattr(st.session_state, 'resultado') and st.session_state.resultado is not None:
        resultado = st.session_state.resultado
        dfs = st.session_state.dfs
        
        if resultado["status"] in ["OPTIMAL", "FEASIBLE"]:
            st.success(f"✅ Horario generado ({resultado['status']})")
            
            enviar_api = st.button("📤 Enviar Horario a API", key="enviar_horario")
            if enviar_api:
                
                for clase in resultado["horario_generado"]:
                    clase_data = {
                        "grupo": clase['grupo'],
                        "dia_semana": clase['dia_semana'],
                        "hora_inicio": clase['hora_inicio'],
                        "hora_fin": clase['hora_fin'],
                        "alumnos": int(clase['alumnos']),
                        "materia_id": int(clase['materia_id']),
                        "salon_id": int(clase['salon_id']),
                        "profesor_id": int(clase['profesor_id'])
                    }
                    
                    try:
                        response = requests.post(
                            f"{BASE_URL}/clases",
                            json=clase_data,
                            headers={'Content-Type': 'application/json'}
                        )
                        
                        if response.status_code in [200, 201]:  # Agregamos 201 como éxito
                            st.success(f"✅ Clase {clase['grupo']} enviada exitosamente (ID: {response.json()['id']})")
                        else:
                            st.error(f"❌ Error al enviar clase {clase['grupo']}: {response.text}")
                            
                    except Exception as e:
                        st.error(f"Error de conexión al enviar clase {clase['grupo']}: {str(e)}")
                        break
            df_horario = pd.DataFrame(resultado["horario_generado"])
            df_horario = df_horario.merge(
                dfs['profesores'][['id', 'nombre']],
                left_on='profesor_id',
                right_on='id',
                suffixes=('', '_profesor')
            )
            df_horario = df_horario.merge(
                dfs['materias'][['id', 'nombre']],
                left_on='materia_id',
                right_on='id',
                suffixes=('', '_materia')
            )
            
            tab1, tab2, tab3 = st.tabs(["Vista por Día", "Estadísticas", "Exportar"])
            
            with tab1:
                dias = sorted(df_horario['dia_semana'].unique())
                for dia in dias:
                    with st.expander(f"📅 {dia}", expanded=True):
                        df_dia = df_horario[df_horario['dia_semana'] == dia].sort_values('hora_inicio')
                        st.dataframe(
                            df_dia[[
                                'grupo', 'hora_inicio', 'hora_fin',
                                'nombre_materia', 'nombre', 'alumnos'
                            ]].style.background_gradient(cmap='Blues'),
                            hide_index=True,
                            use_container_width=True
                        )
            
            with tab2:
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total de clases", len(df_horario))
                with col2:
                    st.metric("Profesores asignados", df_horario['profesor_id'].nunique())
                with col3:
                    st.metric("Materias programadas", df_horario['materia_id'].nunique())
                
                col_charts1, col_charts2 = st.columns(2)
                
                with col_charts1:
                    fig_carga = px.bar(
                        df_horario.groupby('nombre')['grupo'].count().reset_index(),
                        x='nombre',
                        y='grupo',
                        title='Carga por Profesor',
                        labels={'grupo': 'Número de clases', 'nombre': 'Profesor'}
                    )
                    st.plotly_chart(fig_carga, use_container_width=True)
                
                with col_charts2:
                    fig_materias = px.pie(
                        df_horario.groupby('nombre_materia')['grupo'].count().reset_index(),
                        values='grupo',
                        names='nombre_materia',
                        title='Distribución de Materias'
                    )
                    st.plotly_chart(fig_materias, use_container_width=True)
            
            with tab3:
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_horario_export = df_horario.copy()
                    df_horario_export.columns = [col.replace('_', ' ').title() for col in df_horario_export.columns]
                    df_horario_export.to_excel(
                        writer,
                        sheet_name='Horario_Completo',
                        index=False
                    )
                    
                    resumen_prof = df_horario.groupby('nombre').agg({
                        'grupo': 'count',
                        'alumnos': 'sum'
                    }).reset_index()
                    resumen_prof.columns = ['Profesor', 'Total Clases', 'Total Alumnos']
                    resumen_prof.to_excel(
                        writer,
                        sheet_name='Resumen_Profesores',
                        index=False
                    )
                    
                    workbook = writer.book
                    for sheet in workbook.sheetnames:
                        worksheet = workbook[sheet]
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
                
                st.download_button(
                    label="⬇️ Descargar el Horario (Excel)",
                    data=output.getvalue(),
                    file_name=f"horario_generado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if resultado["warnings"]:
                with st.expander("⚠️ Advertencias", expanded=True):
                    for warning in resultado["warnings"]:
                        st.warning(warning)
        
        elif resultado["status"] == "ERROR":
            st.error("❌ No se pudo generar el horario")
            if resultado["errors"]:
                for error in resultado["errors"]:
                    st.error(error)
                

def show_analysis(optimizer):
    st.header("📈 Análisis de Resultados y Métricas")
    
    if not optimizer.is_fitted:
        st.warning("⚠️ No hay un modelo entrenado para analizar. Por favor, entrene el modelo primero.")
        return
    
    # Crear pestañas principales para diferentes tipos de análisis
    tab_perf, tab_dist, tab_patterns, tab_compare = st.tabs([
        "Rendimiento del Modelo",
        "Distribución de Carga",
        "Patrones y Tendencias",
        "Análisis Comparativo"
    ])
    
    with tab_dist:
        st.subheader("📊 Distribución de Carga y Recursos")
        
        # Métricas principales en cards
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric(
                "Precisión del Modelo",
                f"{optimizer.best_model_score:.2%}",
                delta=f"{0.05:.2%}",  # Ejemplo de cambio respecto a la versión anterior
                help="Precisión general del modelo en el conjunto de prueba"
            )
        with col2:
            st.metric(
                "Horarios Generados",
                len(optimizer.performance_history),
                help="Número total de horarios generados exitosamente"
            )
        with col3:
            if hasattr(optimizer.adaptive_agent, 'success_rate_history') and optimizer.adaptive_agent.success_rate_history:
                success_rate = optimizer.adaptive_agent.success_rate_history[-1]
                st.metric(
                    "Tasa de Éxito",
                    f"{success_rate:.2%}",
                    help="Porcentaje de horarios generados sin conflictos"
                )
        with col4:
            if hasattr(optimizer.adaptive_agent, 'learning_rate'):
                st.metric(
                    "Tasa de Aprendizaje",
                    f"{optimizer.adaptive_agent.learning_rate:.3f}",
                    help="Tasa actual de aprendizaje del agente adaptativo"
                )
        
        # Gráfico de evolución del rendimiento
        if optimizer.performance_history:
            st.subheader("📈 Evolución del Rendimiento")
            perf_df = pd.DataFrame(optimizer.performance_history)
            fig_perf = px.line(
                perf_df,
                x='date',
                y='score',
                title='Evolución de la Precisión del Modelo',
                labels={'date': 'Fecha', 'score': 'Precisión'},
                line_shape='spline'
            )
            fig_perf.update_traces(mode='lines+markers')
            st.plotly_chart(fig_perf, use_container_width=True)
            
            # Análisis de tendencia
            if len(perf_df) > 1:
                trend = np.polyfit(range(len(perf_df)), perf_df['score'], 1)[0]
                trend_direction = "positiva" if trend > 0 else "negativa"
                st.info(f"📊 La tendencia general es {trend_direction} con una pendiente de {abs(trend):.4f}")
    
    with tab_dist:
        st.subheader("📊 Distribución de Carga y Recursos")
        
        # Cargar datos actuales
        data = {
            'profesores': optimizer.get_data('profesores'),
            'materias': optimizer.get_data('materias'),
            'salones': optimizer.get_data('salones'),
            'horarios_disponibles': optimizer.get_data('horarios_disponibles'),
            'profesor_materia': optimizer.get_data('profesor_materia')
        }
        
        if all(data.values()):
            dfs = {k: pd.DataFrame(v) for k, v in data.items()}
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Distribución de carga docente
                prof_carga = dfs['profesor_materia'].groupby('profesor_id').size().reset_index()
                prof_carga.columns = ['profesor_id', 'carga']
                prof_carga = prof_carga.merge(dfs['profesores'][['id', 'nombre']], left_on='profesor_id', right_on='id')
                
                fig_carga = px.bar(
                    prof_carga,
                    x='nombre',
                    y='carga',
                    title='Distribución de Carga Docente',
                    labels={'nombre': 'Profesor', 'carga': 'Número de Materias'}
                )
                st.plotly_chart(fig_carga, use_container_width=True)
            
            with col2:
                # Utilización de salones
                salon_stats = dfs['salones'].copy()
                salon_stats['utilization'] = np.random.uniform(0.6, 0.9, len(salon_stats))  # Ejemplo
                
                fig_salones = px.bar(
                    salon_stats,
                    x='codigo',  # Cambiado de 'nombre' a 'codigo'
                    y='utilization',
                    title='Utilización de Salones',
                    labels={'codigo': 'Salón', 'utilization': 'Porcentaje de Utilización'}
                )
                fig_salones.update_traces(marker_color='rgb(55, 83, 109)')
                st.plotly_chart(fig_salones, use_container_width=True)
            
            # Mapa de calor de disponibilidad
            st.subheader("🗓️ Mapa de Calor de Disponibilidad")
            disponibilidad = pd.pivot_table(
                dfs['horarios_disponibles'],
                values='profesor_id',
                index='dia',
                columns='hora_inicio',
                aggfunc='count'
            )
            
            fig_heatmap = px.imshow(
                disponibilidad,
                title='Disponibilidad por Día y Hora',
                labels=dict(x="Hora", y="Día", color="Profesores Disponibles")
            )
            st.plotly_chart(fig_heatmap, use_container_width=True)
    
    with tab_patterns:
        st.subheader("🔍 Análisis de Patrones y Tendencias")
        
        if hasattr(optimizer.adaptive_agent, 'pattern_memory'):
            pattern_data = optimizer.adaptive_agent.pattern_memory
            if pattern_data:
                # Convertir patrones a DataFrame para análisis
                patterns_df = pd.DataFrame([
                    {
                        'pattern': str(k),
                        'success_rate': np.mean(v),
                        'frequency': len(v)
                    }
                    for k, v in pattern_data.items()
                ]).sort_values('success_rate', ascending=False)
                
                # Mostrar patrones más exitosos
                st.subheader("🏆 Patrones Más Exitosos")
                col1, col2 = st.columns(2)
                
                with col1:
                    fig_patterns = px.bar(
                        patterns_df.head(10),
                        x='pattern',
                        y='success_rate',
                        title='Top 10 Patrones por Tasa de Éxito',
                        labels={'pattern': 'Patrón', 'success_rate': 'Tasa de Éxito'}
                    )
                    st.plotly_chart(fig_patterns, use_container_width=True)
                
                with col2:
                    fig_freq = px.scatter(
                        patterns_df,
                        x='frequency',
                        y='success_rate',
                        title='Relación entre Frecuencia y Éxito',
                        labels={'frequency': 'Frecuencia', 'success_rate': 'Tasa de Éxito'}
                    )
                    st.plotly_chart(fig_freq, use_container_width=True)
        
        # Análisis de tendencias temporales
        st.subheader("📅 Tendencias Temporales")
        if optimizer.performance_history:
            temp_df = pd.DataFrame(optimizer.performance_history)
            temp_df['date'] = pd.to_datetime(temp_df['date'])
            temp_df.set_index('date', inplace=True)
            
            # Análisis por hora del día
            temp_df['hour'] = temp_df.index.hour
            hourly_performance = temp_df.groupby('hour')['score'].mean()
            
            fig_hourly = px.line(
                hourly_performance,
                title='Rendimiento por Hora del Día',
                labels={'hour': 'Hora', 'value': 'Rendimiento Promedio'}
            )
            st.plotly_chart(fig_hourly, use_container_width=True)
    
    with tab_compare:
        st.subheader("🔄 Analisis Comparativo")
        
        # Comparación de modelos si hay múltiples entrenamientos
        if optimizer.performance_history:
            model_comparison = pd.DataFrame(optimizer.performance_history)
            
            # Comparar rendimiento por tipo de modelo
            fig_model_comp = px.box(
                model_comparison,
                x='model_type',
                y='score',
                title='Comparación de Rendimiento por Tipo de Modelo',
                labels={'model_type': 'Tipo de Modelo', 'score': 'Precisión'}
            )
            st.plotly_chart(fig_model_comp, use_container_width=True)
            
            # Análisis de parámetros
            st.subheader("⚙️ Analisis de Parametros")
            param_analysis = pd.DataFrame([
                {**record['params'], 'score': record['score']}
                for record in optimizer.performance_history
                if 'params' in record
            ])
            
            if not param_analysis.empty:
                for param in param_analysis.columns:
                    if param != 'score':
                        fig_param = px.scatter(
                            param_analysis,
                            x=param,
                            y='score',
                            title=f'Impacto de {param} en el Rendimiento',
                            trendline="ols"
                        )
                        st.plotly_chart(fig_param, use_container_width=True)
        
        # Métricas de rendimiento del sistema
        st.subheader("⚡ Métricas del Sistema")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric(
                "Tiempo Promedio de Generación",
                "2.5s",  # Ejemplo
                help="Tiempo promedio para generar un horario completo"
            )
        with col2:
            st.metric(
                "Uso de Memoria",
                "256MB",  # Ejemplo
                help="Uso promedio de memoria durante la generación"
            )
        with col3:
            st.metric(
                "Conflictos Resueltos",
                "95%",  # Ejemplo
                help="Porcentaje de conflictos resueltos automáticamente"
            )
    
     # Botón para exportar análisis
    if st.button("📊 Exportar Analisis Completo", use_container_width=True):
        try:
            # Crear un Excel con todos los análisis
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Inicializar un diccionario para almacenar todos los DataFrames a exportar
                export_data = {}
                
                # Rendimiento del modelo
                if optimizer.performance_history:
                    perf_df = pd.DataFrame(optimizer.performance_history)
                    if not perf_df.empty:
                        export_data['Rendimiento'] = perf_df
                
                # Datos actuales del sistema
                data = {
                    'profesores': optimizer.get_data('profesores'),
                    'materias': optimizer.get_data('materias'),
                    'salones': optimizer.get_data('salones'),
                    'horarios_disponibles': optimizer.get_data('horarios_disponibles'),
                    'profesor_materia': optimizer.get_data('profesor_materia')
                }
                
                if all(data.values()):
                    dfs = {k: pd.DataFrame(v) for k, v in data.items()}
                    
                    # Análisis de carga docente
                    prof_carga = dfs['profesor_materia'].groupby('profesor_id').size().reset_index()
                    prof_carga.columns = ['profesor_id', 'carga']
                    prof_carga = prof_carga.merge(
                        dfs['profesores'][['id', 'nombre']], 
                        left_on='profesor_id', 
                        right_on='id'
                    )
                    if not prof_carga.empty:
                        export_data['Carga_Docente'] = prof_carga
                    
                    # Análisis de salones
                    salon_stats = dfs['salones'].copy()
                    salon_stats['utilization'] = np.random.uniform(0.6, 0.9, len(salon_stats))
                    if not salon_stats.empty:
                        export_data['Estadisticas_Salones'] = salon_stats
                    
                    # Análisis de disponibilidad
                    disponibilidad = pd.pivot_table(
                        dfs['horarios_disponibles'],
                        values='profesor_id',
                        index='dia',
                        columns='hora_inicio',
                        aggfunc='count'
                    )
                    if not disponibilidad.empty:
                        export_data['Disponibilidad'] = disponibilidad
                
                # Patrones si existen
                if hasattr(optimizer.adaptive_agent, 'pattern_memory'):
                    pattern_data = optimizer.adaptive_agent.pattern_memory
                    if pattern_data:
                        patterns_df = pd.DataFrame([
                            {
                                'pattern': str(k),
                                'success_rate': np.mean(v),
                                'frequency': len(v)
                            }
                            for k, v in pattern_data.items()
                        ]).sort_values('success_rate', ascending=False)
                        if not patterns_df.empty:
                            export_data['Patrones'] = patterns_df
                
                # Resumen general
                summary_data = {
                    'Métrica': [
                        'Precisión del Modelo',
                        'Total Horarios Generados',
                        'Tasa de Éxito',
                        'Fecha de Análisis'
                    ],
                    'Valor': [
                        f"{getattr(optimizer, 'best_model_score', 0):.2%}",
                        len(optimizer.performance_history),
                        f"{getattr(optimizer.adaptive_agent, 'success_rate_history', [0])[-1]:.2%}" if hasattr(optimizer.adaptive_agent, 'success_rate_history') and optimizer.adaptive_agent.success_rate_history else "N/A",
                        datetime.now().strftime("%Y-%m-%d %H:%M")
                    ]
                }
                export_data['Resumen'] = pd.DataFrame(summary_data)
                
                # Si no hay datos para exportar, crear al menos una hoja con información básica
                if not export_data:
                    export_data['Info'] = pd.DataFrame({
                        'Información': ['No hay datos suficientes para el análisis'],
                        'Fecha': [datetime.now().strftime("%Y-%m-%d %H:%M")]
                    })
                
                # Exportar todos los DataFrames
                for sheet_name, df in export_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Ajustar el ancho de las columnas
                    worksheet = writer.sheets[sheet_name]
                    for idx, col in enumerate(df.columns):
                        max_length = max(
                            df[col].astype(str).apply(len).max(),
                            len(str(col))
                        )
                        worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = max_length + 2
            
            # Botón de descarga
            st.download_button(
                label="⬇️ Descargar Analisis (Excel)",
                data=output.getvalue(),
                file_name=f"analisis_horarios_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.success("✅ Analisis exportado exitosamente")
            
        except Exception as e:
            st.error(f"Error al exportar el análisis: {str(e)}")
            st.error("Detalles del error para depuración:")
            st.exception(e)


if __name__ == "__main__":
    main()
